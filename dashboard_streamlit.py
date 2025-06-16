import streamlit as st
import pandas as pd
from datetime import datetime
import io
# openpyxl é necessário para a formatação do Excel. Instale com: pip install openpyxl
import openpyxl
from openpyxl.styles import PatternFill, Font

# Define 'hoje' no início, sem o componente de hora, para consistência nos cálculos.
hoje = pd.to_datetime(datetime.now().date())

st.set_page_config(page_title="RELATÓRIO DINÂMICO PCP", layout="wide")
st.title("RELATÓRIO DINÂMICO PCP")

URL_CSV = "https://docs.google.com/spreadsheets/d/14i9CpKM87PRXvLAVo_0VUcjNxfxnCLwhhmg0mh5ywkI/gviz/tq?tqx=out:csv&gid=1319025913"

@st.cache_data(ttl=3600)
def carregar_dados():
    """
    Carrega os dados, converte as colunas de data, calcula os dias
    e pré-ordena o DataFrame.
    """
    df = pd.read_csv(URL_CSV)
    # Limpa espaços em branco que possam existir nos nomes das colunas
    df = df.rename(columns=lambda x: x.strip())

    # Converte as colunas de data para datetime. Erros na conversão viram NaT (Not a Time).
    df["DATA ENTREGA PRIMEIRA VALIDACAO"] = pd.to_datetime(df["DATA ENTREGA PRIMEIRA VALIDACAO"], errors="coerce", dayfirst=True)
    df["DATA ALTERACAO STATUS"] = pd.to_datetime(df["DATA ALTERACAO STATUS"], errors="coerce", dayfirst=True)

    # Calcula a diferença em dias usando as colunas datetime.
    # Um valor negativo aqui indica que a data na planilha é futura.
    df["DIAS EM VC"] = (hoje - df["DATA ENTREGA PRIMEIRA VALIDACAO"]).dt.days.astype("Int64")
    df["DIAS ALTERACAO STATUS"] = (hoje - df["DATA ALTERACAO STATUS"]).dt.days.astype("Int64")

    # Pré-ordena os dados: do mais antigo (maior número de dias) para o mais novo.
    df = df.sort_values(
        by=["DIAS EM VC", "DIAS ALTERACAO STATUS"],
        ascending=[False, False],
        na_position='last'
    )

    return df

def format_excel(df):
    """
    Formata o DataFrame para um arquivo Excel com cores condicionais e o retorna como bytes.
    """
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df_to_export = df.copy()
    # Remove a quebra de linha para o nome da coluna no Excel
    df_to_export.rename(columns={'DIAS\nALT.': 'DIAS ALT.', 'PRIM.\nENTREGA': 'PRIM. ENTREGA', 'ALT.\nSTATUS': 'ALT. STATUS'}, inplace=True)

    df_to_export.to_excel(writer, index=False, sheet_name='Relatorio')
    workbook = writer.book
    worksheet = writer.sheets['Relatorio']

    # --- Define Fills (Cores de preenchimento) ---
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    blue_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    white_font = Font(color="FFFFFF")

    # --- Obtém as letras das colunas para formatação ---
    headers = [cell.value for cell in worksheet[1]]
    try:
        # Colunas com os valores numéricos para a lógica
        dias_vc_col_idx = headers.index('DIAS EM VC') + 1
        dias_alt_col_idx = headers.index('DIAS ALT.') + 1
        # Colunas a serem coloridas
        ent_col_idx = headers.index('ENT') + 1
        alt_col_idx = headers.index('ALT') + 1
    except ValueError:
        # Se alguma coluna não for encontrada, retorna o Excel sem formatação de cor
        writer.close()
        return output.getvalue()

    # Itera sobre as linhas da planilha para aplicar a formatação
    for row in range(2, worksheet.max_row + 1):
        # --- Formatação para a coluna 'ENT' ---
        dias_vc_cell = worksheet.cell(row=row, column=dias_vc_col_idx)
        if pd.notna(dias_vc_cell.value):
            dias = int(dias_vc_cell.value)
            cell_to_format = worksheet.cell(row=row, column=ent_col_idx)
            if dias <= 21:
                cell_to_format.fill = green_fill
            elif dias <= 30:
                cell_to_format.fill = orange_fill
            else: # > 30
                cell_to_format.fill = light_red_fill

        # --- Formatação para a coluna 'ALT' ---
        dias_alt_cell = worksheet.cell(row=row, column=dias_alt_col_idx)
        if pd.notna(dias_alt_cell.value):
            dias = int(dias_alt_cell.value)
            cell_to_format = worksheet.cell(row=row, column=alt_col_idx)
            if dias <= 7: cell_to_format.fill = green_fill
            elif dias <= 14: cell_to_format.fill = blue_fill
            elif dias <= 21: cell_to_format.fill = yellow_fill
            elif dias <= 30: cell_to_format.fill = orange_fill
            elif dias <= 45: cell_to_format.fill = light_red_fill
            else: # > 45
                cell_to_format.fill = black_fill
                cell_to_format.font = white_font

    # --- Auto-ajuste da largura das colunas ---
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    writer.close()
    return output.getvalue()


df = carregar_dados()

# --- Filtros no topo ---
st.markdown("## Filtros")
col1, col2 = st.columns(2)
with col1:
    clientes = st.multiselect("Cliente", sorted(df["CLIENTE"].dropna().unique()), default=None)
    gps = st.multiselect("GP", sorted(df["GP"].dropna().unique()), default=None)
with col2:
    produtos = st.multiselect("Produto", sorted(df["PRODUTO"].dropna().unique()), default=None)
    status = st.multiselect("Status", sorted(df["STATUS"].dropna().unique()), default=["VC", "VC R1", "VC R2", "VC ADD"])

# Tratamento para evitar erro se a coluna de data estiver vazia ou sem datas válidas
if not df["DATA ENTREGA PRIMEIRA VALIDACAO"].dropna().empty:
    data_min = df["DATA ENTREGA PRIMEIRA VALIDACAO"].min()
    data_max = df["DATA ENTREGA PRIMEIRA VALIDACAO"].max()

    if pd.notna(data_min) and pd.notna(data_max):
        col3, col4 = st.columns(2)
        with col3:
            data_inicial = st.date_input("Data Inicial", value=data_min, min_value=data_min, max_value=data_max, format="DD/MM/YYYY")
        with col4:
            data_final = st.date_input("Data Final", value=data_max, min_value=data_min, max_value=data_max, format="DD/MM/YYYY")
    else:
        st.warning("Não foi possível determinar o intervalo de datas para o filtro.")
        data_inicial, data_final = None, None
else:
    st.warning("Nenhuma data de entrega válida encontrada para definir os filtros.")
    data_inicial, data_final = None, None


# --- Aplicar filtros ---
df_filtrado = df.copy()
if clientes:
    df_filtrado = df_filtrado[df_filtrado["CLIENTE"].isin(clientes)]
if gps:
    df_filtrado = df_filtrado[df_filtrado["GP"].isin(gps)]
if produtos:
    df_filtrado = df_filtrado[df_filtrado["PRODUTO"].isin(produtos)]
if status:
    df_filtrado = df_filtrado[df_filtrado["STATUS"].isin(status)]

# Filtro de data, garantindo que as datas não sejam nulas antes de comparar
if data_inicial and data_final:
    df_filtrado = df_filtrado[
        df_filtrado["DATA ENTREGA PRIMEIRA VALIDACAO"].notna() &
        (df_filtrado["DATA ENTREGA PRIMEIRA VALIDACAO"] >= pd.to_datetime(data_inicial)) &
        (df_filtrado["DATA ENTREGA PRIMEIRA VALIDACAO"] <= pd.to_datetime(data_final))
    ]

# --- Funções para indicadores visuais ---
def gerar_bolinha_status(dias):
    if pd.isna(dias):
        return "⚪"
    dias = int(dias)
    if dias < 0: return "📅" # Indica data futura
    if dias <= 7: return "🟢"
    elif dias <= 14: return "🔵"
    elif dias <= 21: return "🟡"
    elif dias <= 30: return "🟠"
    elif dias <= 45: return "🔴"
    else: return "⚫"

def gerar_bolinha_vc(dias):
    if pd.isna(dias):
        return "⚪"
    dias = int(dias)
    if dias < 0: return "📅" # Emoji de calendário para indicar data futura
    if dias <= 21: return "🟢"
    elif dias <= 30: return "🟠"
    else: return "🔴"

# --- Legendas ---
st.markdown("---")
st.markdown("#### Legendas dos Indicadores")
st.markdown(
    """
    **Entrega (ENT):** &nbsp; 🟢 Até 21 dias &nbsp; | &nbsp; 🟠 22 a 30 dias &nbsp; | &nbsp; 🔴 Acima de 30 dias <br>
    **Alteração (ALT):** &nbsp; 🟢 Até 7 &nbsp; | &nbsp; 🔵 8 a 14 &nbsp; | &nbsp; 🟡 15 a 21 &nbsp; | &nbsp; 🟠 22 a 30 &nbsp; | &nbsp; 🔴 31 a 45 &nbsp; | &nbsp; ⚫ Acima de 45 dias <br>
    **Geral:** &nbsp; 📅 Data Futura &nbsp; | &nbsp; ⚪ Dado Ausente
    """,
    unsafe_allow_html=True
)
st.markdown("---")


# --- Geração de colunas visuais e exibição da tabela ---
# Aplica as funções para gerar os indicadores visuais no DataFrame já filtrado
df_filtrado["STATUS VISUAL"] = df_filtrado["DIAS ALTERACAO STATUS"].apply(gerar_bolinha_status)
df_filtrado["VC VISUAL"] = df_filtrado["DIAS EM VC"].apply(gerar_bolinha_vc)

# Cria as colunas de data formatadas como string APENAS para exibição
df_filtrado['DATA PRIMEIRA ENTREGA'] = df_filtrado['DATA ENTREGA PRIMEIRA VALIDACAO'].dt.strftime('%d/%m/%Y')
df_filtrado['DATA ALTERACAO STATUS_STR'] = df_filtrado['DATA ALTERACAO STATUS'].dt.strftime('%d/%m/%Y')

# A ordenação inicial já foi feita na função carregar_dados()
# Apenas selecionamos e renomeamos as colunas para exibição
tabela_para_exibir = df_filtrado[[
    'CLIENTE', 'PROJETO', 'STATUS',
    'DATA PRIMEIRA ENTREGA', 'DIAS EM VC', 'VC VISUAL',
    'DATA ALTERACAO STATUS_STR', 'DIAS ALTERACAO STATUS', 'STATUS VISUAL',
    'OBSERVACOES'
]]

# Renomeia as colunas para o formato final de exibição com os novos títulos
tabela_para_exibir = tabela_para_exibir.rename(columns={
    'VC VISUAL': 'ENT',
    'STATUS VISUAL': 'ALT',
    'DATA PRIMEIRA ENTREGA': 'PRIM.\nENTREGA',
    'DATA ALTERACAO STATUS_STR': 'ALT.\nSTATUS',
    'DIAS ALTERACAO STATUS': 'DIAS\nALT.'
})

# --- Exibição da Tabela e Botão de Download ---
col_titulo, col_botao = st.columns([0.8, 0.2])

with col_titulo:
    st.markdown("### Tabela Detalhada com Indicadores Visuais")

with col_botao:
    if not tabela_para_exibir.empty:
        # Prepara uma versão do dataframe para exportação
        df_export = tabela_para_exibir.copy()
        excel_data = format_excel(df_export)
        st.download_button(
            label="📥 Exportar para Excel",
            data=excel_data,
            file_name=f"relatorio_pcp_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# Usa st.dataframe para uma tabela interativa que permite ordenação pelo cabeçalho
st.dataframe(
    tabela_para_exibir,
    use_container_width=True,
    hide_index=True,
    height=735 # (20 linhas * 35px por linha) + 35px para o cabeçalho
)
